#-*- coding: utf-8 -*-
"""
Lógica de backend para análise de aulas e-Pratika.
Este arquivo é projetado para ser importado como um módulo.
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# === FUNÇÕES DE LÓGICA ===

def iniciar_driver():
    """Inicializa e retorna uma instância do WebDriver do Chrome."""
    options = Options()
    options.page_load_strategy = 'eager'  # Não esperar o carregamento completo da página
    # options.add_argument('--headless') # Descomente para rodar sem abrir a janela do navegador
    options.add_argument('--disable-gpu')
    options.add_argument("--window-size=1366,768")
    return webdriver.Chrome(options=options)

def fazer_login(driver, login_url, local_id, usuario, senha, progress_callback):
    """Executa o processo de login no site."""
    progress_callback(f"Navegando para a URL de login...\n")
    driver.get(login_url)

    try:
        WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.ID, 'local')))
        progress_callback("Página de login carregada. Preenchendo dados...\n")
        Select(driver.find_element(By.ID, 'local')).select_by_value(local_id)
        driver.find_element(By.NAME, 'usuario').send_keys(usuario)
        driver.find_element(By.NAME, 'senha').send_keys(senha)
        driver.find_element(By.ID, 'entrar').click()
        WebDriverWait(driver, 20).until(EC.url_changes(login_url))
        progress_callback(f"Login realizado com sucesso! URL atual: {driver.current_url}\n")
    except TimeoutException:
        progress_callback("Erro: Timeout ao tentar fazer login. Verifique a URL e a conexão.\n")
        raise

def obter_registros_de_pagina(driver, url, progress_callback):
    """Obtém a lista de registros de uma única página de relatório."""
    progress_callback(f"Acessando página: {url}\n")
    driver.get(url)
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table.tabela')))
    except TimeoutException:
        progress_callback("Nenhum registro encontrado nesta página (timeout). Fim da paginação.\n")
        return []

    linhas = driver.find_elements(By.CSS_SELECTOR, 'table.tabela tbody tr')
    if not linhas or (len(linhas) == 1 and "Nenhum registro encontrado" in linhas[0].text):
        return []

    dados = []
    for linha in linhas:
        colunas = linha.find_elements(By.TAG_NAME, 'td')
        if len(colunas) >= 19:
            try:
                link_acao = colunas[18].find_element(By.TAG_NAME, 'a').get_attribute('href')
                dados.append({
                    'Data Hora Início': colunas[0].text.strip(),
                    'Nome': colunas[3].text.strip(),
                    'RENACH': colunas[5].text.strip(),
                    'Instrutor': colunas[7].text.strip(),
                    'Tablet': colunas[10].text.strip(),
                    'Veículo': colunas[8].text.strip(),
                    'Link': link_acao
                })
            except NoSuchElementException:
                # Linha pode não ter um link de ação, ignorar
                pass
    return dados

def extrair_dados_do_link(driver, url, progress_callback):
    """Visita um único link de detalhes e extrai todas as informações de uma vez."""
    progress_callback(f"  Analisando detalhes de: {url[:50]}...\n")
    driver.get(url)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
    except TimeoutException:
        progress_callback(f"  Timeout ao carregar a página de detalhes. Pulando.\n")
        return { 'inicio': 0, 'final': 0, 'percurso': 0, 'trajeto': 0, 'auditoria': "Erro de Timeout" }

    # Contagem de imagens
    qtd_inicio = len(driver.find_elements(By.XPATH, '//div[contains(text(), "Foto Aluno Início")]')) + \
                 len(driver.find_elements(By.XPATH, '//div[contains(text(), "Foto Aluno Auditoria Início")]'))
    
    qtd_final = len(driver.find_elements(By.XPATH, '//div[contains(text(), "Foto Aluno Final")]')) + \
                len(driver.find_elements(By.XPATH, '//div[contains(text(), "Foto Aluno Auditoria Final")]'))

    qtd_percurso = len(driver.find_elements(By.XPATH, '//div[contains(text(), "Foto do Percurso")]'))
    qtd_trajeto = len(driver.find_elements(By.XPATH, '//div[contains(text(), "Imagem Trajeto")]'))

    # Obter auditorias
    auditoria_texto = "Nenhuma"
    try:
        botao_auditoria = driver.find_element(By.XPATH, '//span[text()="Auditorias"]')
        botao_auditoria.click()
        modal = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.ui-dialog')))
        
        try:
            auditoria_texto = modal.find_element(By.CSS_SELECTOR, 'table').text.strip()
        except NoSuchElementException:
            auditoria_texto = modal.text.strip()

        # Tenta fechar o modal
        driver.find_element(By.CSS_SELECTOR, '.ui-dialog-titlebar-close').click()
        WebDriverWait(driver, 10).until(EC.invisibility_of_element(modal))

    except (NoSuchElementException, TimeoutException):
        # Se o botão ou modal não existem, apenas continua
        pass
    except Exception as e:
        progress_callback(f"  Erro inesperado ao obter auditorias: {e}\n")

    return {
        'inicio': qtd_inicio,
        'final': qtd_final,
        'percurso': qtd_percurso,
        'trajeto': qtd_trajeto,
        'auditoria': auditoria_texto.replace('\n', '; ')
    }

def salvar_em_txt(dados, caminho, progress_callback):
    """Salva os dados processados em um arquivo de texto formatado (TSV)."""
    with open(caminho, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(['Data Hora Início', 'Nome', 'RENACH', 'Instrutor',
                         'Foto de Início', 'Foto de Fim', 'Foto do Percurso', 'Foto do trajeto',
                         'Tablet', 'Veículo', 'Ação(link)', 'Auditorias'])
        writer.writerows(dados)
    progress_callback(f"Arquivo de resultado salvo em: {caminho}\n")

def salvar_em_excel(dados, caminho, progress_callback):
    """Salva os dados processados em um arquivo Excel (.xlsx) com formatação.
    Os dados devem incluir o cabeçalho como a primeira linha.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Análise de Aulas"

    # Estilos
    header_font = Font(bold=True)
    even_row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Cinza claro
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Branco
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Vermelho claro para alerta

    # Escreve o cabeçalho e aplica estilo
    headers = dados[0] if dados else []
    ws.append(headers)
    for col_idx, header_text in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).font = header_font

    # Escreve os dados e aplica formatação de linhas alternadas e condicional
    for row_idx, row_data in enumerate(dados[1:], 2): # Começa da segunda linha de dados (linha 2 do Excel)
        ws.append(row_data)
        
        # Verifica condição para alerta vermelho
        foto_inicio = row_data[4] if len(row_data) > 4 else '' # 'Foto de Início'
        foto_percurso = row_data[6] if len(row_data) > 6 else '' # 'Foto do Percurso'

        apply_red_alert = False
        if foto_inicio == '0' or foto_percurso == '0':
            apply_red_alert = True

        if apply_red_alert:
            for cell in ws[row_idx]:
                cell.fill = red_fill
        elif row_idx % 2 == 0: # Linhas pares (no Excel, 2, 4, 6...)
            for cell in ws[row_idx]:
                cell.fill = even_row_fill
        else: # Linhas ímpares (no Excel, 3, 5, 7...)
            for cell in ws[row_idx]:
                cell.fill = odd_row_fill

    # Ajusta a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # Adiciona um pouco de padding
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # --- Criação do Dashboard ---
    try:
        _criar_dashboard(wb, dados, progress_callback)
    except Exception as e:
        progress_callback(f"Ocorreu um erro inesperado durante a criação do dashboard: {e}\n")

    wb.save(caminho)
    progress_callback(f"Arquivo Excel salvo em: {caminho}\n")

def _criar_dashboard(workbook, dados, progress_callback):
    """Cria a aba de dashboard com gráficos."""
    if len(dados) <= 1:
        progress_callback("Nenhum dado para gerar o dashboard.\n")
        return

    progress_callback("Criando aba de Dashboard...\n")
    ws_dashboard = workbook.create_sheet("Dashboard")

    # --- Agregação de Dados ---
    headers = dados[0]
    dados_sem_cabecalho = dados[1:]
    
    # Encontrar índices das colunas pelo nome do cabeçalho
    try:
        idx_instrutor = headers.index('Instrutor')
        idx_foto_inicio = headers.index('Foto de Início')
        idx_foto_percurso = headers.index('Foto do Percurso')
    except ValueError as e:
        progress_callback(f"Erro: Cabeçalho esperado não encontrado no arquivo de dados: {e}\n")
        return

    aulas_por_instrutor = {}
    erros_por_instrutor = {}
    tipos_de_erro = {"Falta Foto de Início": 0, "Falta Foto de Percurso": 0, "Falta Ambas": 0}

    for row in dados_sem_cabecalho:
        instrutor = row[idx_instrutor]
        foto_inicio_ok = row[idx_foto_inicio] != '0'
        foto_percurso_ok = row[idx_foto_percurso] != '0'

        # Contagem total de aulas
        aulas_por_instrutor[instrutor] = aulas_por_instrutor.get(instrutor, 0) + 1

        # Contagem de erros
        if not foto_inicio_ok or not foto_percurso_ok:
            erros_por_instrutor[instrutor] = erros_por_instrutor.get(instrutor, 0) + 1
            if not foto_inicio_ok and not foto_percurso_ok:
                tipos_de_erro["Falta Ambas"] += 1
            elif not foto_inicio_ok:
                tipos_de_erro["Falta Foto de Início"] += 1
            else:
                tipos_de_erro["Falta Foto de Percurso"] += 1

    # --- Escrever dados agregados na planilha para os gráficos ---
    
    # Dados para Gráfico 1 e 2: Aulas e Erros por Instrutor
    ws_dashboard.append(["Instrutor", "Total de Aulas", "Aulas com Erro"])
    instrutores = sorted(aulas_por_instrutor.keys())
    for instrutor in instrutores:
        ws_dashboard.append([instrutor, aulas_por_instrutor.get(instrutor, 0), erros_por_instrutor.get(instrutor, 0)])
    
    # Dados para Gráfico 3: Tipos de Erro
    ws_dashboard.append([]) # Linha em branco
    ws_dashboard.append(["Tipo de Erro", "Quantidade"])
    start_row_erros = len(instrutores) + 3
    for tipo, qtd in tipos_de_erro.items():
        ws_dashboard.append([tipo, qtd])

    # --- Criação dos Gráficos ---

    # Gráfico 1: Total de Aulas por Instrutor
    chart1 = BarChart()
    chart1.title = "Total de Aulas por Instrutor"
    chart1.y_axis.title = "Quantidade de Aulas"
    chart1.x_axis.title = "Instrutor"
    
    data1 = Reference(ws_dashboard, min_col=2, min_row=1, max_row=len(instrutores) + 1, max_col=2)
    cats1 = Reference(ws_dashboard, min_col=1, min_row=2, max_row=len(instrutores) + 1)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.legend = None
    ws_dashboard.add_chart(chart1, "A5")

    # Gráfico 2: Aulas com Erro por Instrutor
    chart2 = BarChart()
    chart2.title = "Aulas com Erro por Instrutor"
    chart2.y_axis.title = "Quantidade de Aulas com Erro"
    chart2.x_axis.title = "Instrutor"

    data2 = Reference(ws_dashboard, min_col=3, min_row=1, max_row=len(instrutores) + 1, max_col=3)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats1) # Reutiliza as categorias do gráfico 1
    chart2.legend = None
    ws_dashboard.add_chart(chart2, "J5")

    # Gráfico 3: Tipos de Erro
    chart3 = PieChart()
    chart3.title = "Distribuição de Tipos de Erro"
    
    labels3 = Reference(ws_dashboard, min_col=1, min_row=start_row_erros, max_row=start_row_erros + len(tipos_de_erro) -1)
    data3 = Reference(ws_dashboard, min_col=2, min_row=start_row_erros - 1, max_row=start_row_erros + len(tipos_de_erro) -1)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(labels3)
    ws_dashboard.add_chart(chart3, "A25")

    progress_callback("Dashboard criado com sucesso.\n")


def exportar_para_excel_existente(caminho_excel, progress_callback):
    """
    Lê o arquivo 'resultado_analise.txt' e o exporta para um arquivo Excel.
    """
    progress_callback(f"Tentando exportar dados de 'resultado_analise.txt' para '{caminho_excel}'...\n")
    try:
        with open('resultado_analise.txt', 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f, delimiter='\t')
            dados_analise_final = list(reader)
    except FileNotFoundError:
        progress_callback("Erro: Arquivo 'resultado_analise.txt' não encontrado. Execute uma análise primeiro.\n")
        return

    if not dados_analise_final or len(dados_analise_final) <= 1:
        progress_callback("Nenhum dado para exportar em 'resultado_analise.txt'.\n")
        return

    salvar_em_excel(dados_analise_final, caminho_excel, progress_callback)

def analisar_txt_final(entrada, saida, progress_callback):
    """Executa a análise secundária para agrupar aulas."""
    try:
        with open(entrada, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            linhas = list(reader)
    except FileNotFoundError:
        progress_callback(f"Arquivo de entrada {entrada} não encontrado para análise final.\n")
        return

    if len(linhas) <= 1:
        progress_callback("Nenhum dado para analisar no arquivo de resultado.\n")
        with open(saida, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            writer.writerow(linhas[0] + ["Grupo"] if linhas else ["Grupo"])
        progress_callback(f"Arquivo de análise final salvo em: {saida}\n")
        return

    cabecalho = linhas[0]
    dados_originais = linhas[1:]
    cabecalho.append("Grupo")

    # Estrutura para armazenar os dados processados: [linha_original, objeto_datetime, grupo_str]
    dados_processados = []
    for linha in dados_originais:
        try:
            # Garante que a linha tenha colunas suficientes antes de acessar
            if linha:
                dt = datetime.strptime(linha[0], "%d/%m/%y %H:%M")
                dados_processados.append([linha, dt, "1"]) # Adiciona a linha, o datetime e o grupo padrão
        except (ValueError, IndexError):
            # Ignora linhas com data inválida ou formato incorreto
            pass

    usados = set()
    for i in range(len(dados_processados)):
        if i in usados:
            continue
        
        _, dt_i, _ = dados_processados[i]
        nome_i, renach_i = dados_processados[i][0][1], dados_processados[i][0][2]

        for j in range(i + 1, len(dados_processados)):
            if j in usados:
                continue

            _, dt_j, _ = dados_processados[j]
            nome_j, renach_j = dados_processados[j][0][1], dados_processados[j][0][2]

            if nome_i == nome_j and renach_i == renach_j:
                diff = abs((dt_j - dt_i).total_seconds()) / 60
                if 40 <= diff <= 55:  # Margem de tolerância
                    dados_processados[i][2] = "1-2" # Atualiza o grupo
                    dados_processados[j][2] = "2-2" # Atualiza o grupo
                    usados.add(i)
                    usados.add(j)
                    break 

    # Prepara os dados para salvar, combinando a linha original com o grupo
    dados_formatados = [linha_original + [grupo] for linha_original, _, grupo in dados_processados]

    with open(saida, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(cabecalho)
        writer.writerows(dados_formatados)
    progress_callback(f"Arquivo de análise final salvo em: {saida}\n")


# === FUNÇÃO PRINCIPAL DE EXECUÇÃO ===

def executar_analise_completa(login_url, usuario, senha, local_id, progress_callback, filters=None):
    """Função principal que orquestra todo o processo de análise."""
    driver = iniciar_driver()
    try:
        fazer_login(driver, login_url, local_id, usuario, senha, progress_callback)
        
        base_url = login_url
        pagina = 1
        todos_os_registros = []
        while True:
            url_pagina = base_url if pagina == 1 else f"{base_url}/{pagina}"
            registros_pagina = obter_registros_de_pagina(driver, url_pagina, progress_callback)
            if not registros_pagina:
                break
            todos_os_registros.extend(registros_pagina)
            pagina += 1
            if pagina > 50: # Trava de segurança para evitar loop infinito
                progress_callback("Atingido limite de 50 páginas. Interrompendo.\n")
                break

        progress_callback(f"Total de {len(todos_os_registros)} registros encontrados nas páginas.\n")

        if not todos_os_registros:
            progress_callback("Nenhum registro para processar. Encerrando.\n")
            # Cria arquivos vazios com cabeçalho
            salvar_em_txt([], 'resultado_formatado.txt', progress_callback)
            analisar_txt_final('resultado_formatado.txt', 'resultado_analise.txt', progress_callback)
            return

        # --- Aplicar Filtros ---
        if filters:
            progress_callback("Aplicando filtros aos registros...\n")
            registros_filtrados = []
            for registro in todos_os_registros:
                match = True
                # Filtros de texto
                if filters.get('nome') and filters['nome'].lower() not in registro['Nome'].lower():
                    match = False
                if match and filters.get('renach') and filters['renach'].lower() not in registro['RENACH'].lower():
                    match = False
                if match and filters.get('instrutor') and filters['instrutor'].lower() not in registro['Instrutor'].lower():
                    match = False
                if match and filters.get('tablet') and filters['tablet'].lower() not in registro['Tablet'].lower():
                    match = False
                if match and filters.get('veiculo') and filters['veiculo'].lower() not in registro['Veículo'].lower():
                    match = False
                
                # Filtros de data
                data_inicio_registro_str = registro['Data Hora Início'].split(' ')[0] # Pega apenas a data
                try:
                    data_inicio_registro = datetime.strptime(data_inicio_registro_str, "%d/%m/%y").date()
                except ValueError:
                    progress_callback(f"Aviso: Formato de data inválido no registro: {registro['Data Hora Início']}. Ignorando filtro de data para este registro.\n")
                    data_inicio_registro = None

                if data_inicio_registro:
                    if filters.get('data_inicio'):
                        try:
                            filtro_data_inicio = datetime.strptime(filters['data_inicio'], "%d/%m/%y").date()
                            if data_inicio_registro < filtro_data_inicio:
                                match = False
                        except ValueError:
                            progress_callback(f"Aviso: Formato de 'Data Início' inválido no filtro: {filters['data_inicio']}. Ignorando este filtro.\n")

                    if match and filters.get('data_fim'):
                        try:
                            filtro_data_fim = datetime.strptime(filters['data_fim'], "%d/%m/%y").date()
                            if data_inicio_registro > filtro_data_fim:
                                match = False
                        except ValueError:
                            progress_callback(f"Aviso: Formato de 'Data Fim' inválido no filtro: {filters['data_fim']}. Ignorando este filtro.\n")

                if match:
                    registros_filtrados.append(registro)
            todos_os_registros = registros_filtrados
            progress_callback(f"Total de {len(todos_os_registros)} registros após a aplicação dos filtros.\n")
        # --- Fim da Aplicação de Filtros ---

        if not todos_os_registros:
            progress_callback("Nenhum registro para processar após a filtragem. Encerrando.\n")
            salvar_em_txt([], 'resultado_formatado.txt', progress_callback)
            analisar_txt_final('resultado_formatado.txt', 'resultado_analise.txt', progress_callback)
            return

        # Processamento otimizado
        dados_completos = []
        total = len(todos_os_registros)
        for i, r in enumerate(todos_os_registros):
            # progress_callback(f"Processando registro {i+1}/{total}...\n") # Removido para reduzir verbosidade
            dados_link = extrair_dados_do_link(driver, r['Link'], progress_callback)
            
            linha = [
                r['Data Hora Início'], r['Nome'], r['RENACH'], r['Instrutor'],
                str(dados_link['inicio']), str(dados_link['final']), 
                str(dados_link['percurso']), str(dados_link['trajeto']),
                r['Tablet'], r['Veículo'], r['Link'], dados_link['auditoria']
            ]
            dados_completos.append(linha)

        # Salvar e analisar
        salvar_em_txt(dados_completos, 'resultado_formatado.txt', progress_callback)
        analisar_txt_final('resultado_formatado.txt', 'resultado_analise.txt', progress_callback)

        progress_callback("\nAnálise concluída com sucesso!\n")

    except Exception as e:
        progress_callback(f"\nOcorreu um erro crítico: {e}\n")
    finally:
        driver.quit()
        progress_callback("Navegador fechado.\n")